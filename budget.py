#!/usr/bin/env python3
import openpyxl as opx
import datetime
import upload
import analysis

budget = opx.load_workbook(analysis.budget)
records = budget.worksheets[0]
today = datetime.datetime.now().strftime("%d/%m/%Y")  # date
choice = 2
#print(today)


def rec_income(sign):
    item = str(input("Enter Name Of Item: "))
    value = sign * float(input("\nEnter Amount: "))
    category = str(input("\nEnter Name Of Category: "))
    comment = str(input("\nDescription: "))

    records.cell(row=Row, column=1).value = today
    records.cell(row=Row, column=2).value = item
    records.cell(row=Row, column=3).value = value
    records.cell(row=Row, column=4).value = category
    records.cell(row=Row, column=5).value = comment


#print(len(records['A']))

while choice != 0:
    print("\n=================================================================\n")
    print("\n=================================================================\n\n")
    Row = int(len(records['A']) + 1)
    choice = int(
        input("\t[1] - Expenditure\n\t[2] - Income\n\n\t[0] - Exit\n\t\t"))
    if choice != 0:
        if choice == 1:
            choice = int(-1)
        if choice == 2:
            choice = 1

        rec_income(choice)

    budget.save(analysis.budget)