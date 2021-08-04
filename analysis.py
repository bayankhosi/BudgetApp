import datetime
import openpyxl as opx
import pandas as pd


budget = opx.load_workbook('./Files/budget.xlsx')
records = budget.worksheets[0]
month = int(datetime.datetime.now().strftime("%m"))  # month number


class month:

    def monthly_total():    # monthly data
        ro = int(input("""
        Enter Number of Month you wanna view
        (eg. May = 5)
        """))

        income_list = []
        income_type_list = []

        for row in records.iter_rows(min_row=5, max_row=7):     # income

            cash = row[ro].value        # the cash
            income_list.append(cash)

            income_type = row[0].value
            income_type_list.append(income_type)

        income_dict = {'Amount': income_list}

        income_dataframe = pd.DataFrame(income_dict, index=income_type_list)

        print(income_dataframe.transpose())

        home_list = []

        for row in records.iter_rows(min_row=12, max_row=16):     # home expenses
            cash = row[ro].value        # the cash
            home_list.append(cash)

            home_type = row[0].value

        daily_list = []

        for row in records.iter_rows(min_row=20, max_row=25):     # daily expenses
            cash = row[ro].value        # the cash
            daily_list.append(cash)

            daily_type = row[0].value

        trans_list = []

        for row in records.iter_rows(min_row=29, max_row=34):     # trans expenses
            cash = row[ro].value        # the cash
            trans_list.append(cash)

            trans_type = row[0].value

        entertain_list = []


        for row in records.iter_rows(min_row=38, max_row=41):     # entertainment expenses
            cash = row[ro].value        # the cash
            entertain_list.append(cash)

            entertain_type = row[0].value

        health_list = []

        for row in records.iter_rows(min_row=45, max_row=51):     # health expenses
            cash = row[ro].value        # the cash
            health_list.append(cash)

            health_type = row[0].value

        vaca_list = []

        for row in records.iter_rows(min_row=55, max_row=60):     # vacation expenses
            cash = row[ro].value        # the cash
            vaca_list.append(cash)

            vaca_type = row[0].value

        rec_list = []


        for row in records.iter_rows(min_row=64, max_row=67):     # recreation expenses
            cash = row[ro].value        # the cash
            rec_list.append(cash)

            rec_type = row[0].value

        sub_list = []


        for row in records.iter_rows(min_row=71, max_row=77):     # subscriptions expenses
            cash = row[ro].value        # the cash
            sub_list.append(cash)

            sub_type = row[0].value

        pers_list = []

        for row in records.iter_rows(min_row=81, max_row=85):     # personal expenses
            cash = row[ro].value        # the cash
            pers_list.append(cash)

            pers_type = row[0].value

        ob_list = []

        for row in records.iter_rows(min_row=89, max_row=93):     # obligations expenses
            cash = row[ro].value        # the cash
            ob_list.append(cash)

            ob_type = row[0].value

        misc_list = []

        for row in records.iter_rows(min_row=97, max_row=101):     # misc expenses
            cash = row[ro].value        # the cash
            misc_list.append(cash)

            misc_type = row[0].value

        total_list = []

        for row in records.iter_rows(min_row=105, max_row=106):     # total expenses
            cash = row[ro].value        # the cash
            total_list.append(cash)

            total_type = row[0].value

        expenses_dict = {'Home': sum(home_list),
                         'Daily': sum(daily_list),
                         'Tranportation': sum(trans_list),
                         'Entertainment': sum(entertain_list),
                         'Health': sum(health_list),
                         'Vacation': sum(vaca_list),
                         'Recreation': sum(rec_list),
                         'Subscriptions': sum(sub_list),
                         'Personal': sum(pers_list),
                         'Obligations': sum(ob_list),
                         'Misc': sum(misc_list)
                         }

        exp = pd.DataFrame(expenses_dict, index=['Amount'])

        print('\n', exp, '\n')


    def monthly_category():
        ro = 1 + int(input("""
            Enter Number of Month you wanna view
            (eg. May = 5)
            """))

        income_list = []
        income_type_list = []

        for row in records.iter_rows(min_row=5, max_row=7):     # income

            cash = row[ro].value        # the cash
            income_list.append(cash)

            income_type = row[0].value
            income_type_list.append(income_type)

        income_dict = {'Amount': income_list}

        income_dataframe = pd.DataFrame(income_dict, index=income_type_list)

        print(income_dataframe.transpose())

        home_list = []
        home_type_list = []

        for row in records.iter_rows(min_row=12, max_row=16):     # home expenses
            cash = row[ro].value        # the cash
            home_list.append(cash)

            home_type = row[0].value
            home_type_list.append(home_type)

        home_dict = {'Amount': home_list}
        home_dataframe = pd.DataFrame(home_dict, index=home_type_list)

        daily_list = []
        daily_type_list = []

        for row in records.iter_rows(min_row=20, max_row=25):     # daily expenses
            cash = row[ro].value        # the cash
            daily_list.append(cash)

            daily_type = row[0].value
            daily_type_list.append(home_type)


        daily_dict = {'Amount': daily_list}
        daily_dataframe = pd.DataFrame(daily_dict, index=daily_type_list)

        trans_list = []
        trans_type_list = []

        for row in records.iter_rows(min_row=29, max_row=34):     # trans expenses
            cash = row[ro].value        # the cash
            trans_list.append(cash)

            trans_type = row[0].value
            trans_type_list.append(trans_type)

        trans_dict = {'Amount': trans_list}
        trans_dataframe = pd.DataFrame(trans_dict, index=trans_type_list)

        entertain_list = []
        entertain_type_list = []

        for row in records.iter_rows(min_row=38, max_row=41):     # entertainment expenses
            cash = row[ro].value        # the cash
            entertain_list.append(cash)

            entertain_type = row[0].value
            entertain_type_list.append(entertain_type)

        entertain_dict = {'Amount': entertain_list}
        entertain_dataframe = pd.DataFrame(entertain_dict, index=entertain_type_list)

        health_list = []
        health_type_list = []

        for row in records.iter_rows(min_row=45, max_row=51):     # health expenses
            cash = row[ro].value        # the cash
            health_list.append(cash)

            health_type = row[0].value
            health_type_list.append(health_type)

        health_dict = {'Amount': health_list}
        health_dataframe = pd.DataFrame(health_dict, index=health_type_list)

        vaca_list = []
        vaca_type_list = []

        for row in records.iter_rows(min_row=55, max_row=60):     # vacation expenses
            cash = row[ro].value        # the cash
            vaca_list.append(cash)

            vaca_type = row[0].value
            vaca_type_list.append(vaca_type)

        vaca_dict = {'Amount': vaca_list}
        vaca_dataframe = pd.DataFrame(vaca_dict, index=vaca_type_list)

        rec_list = []
        rec_type_list = []


        for row in records.iter_rows(min_row=64, max_row=67):     # recreation expenses
            cash = row[ro].value        # the cash
            rec_list.append(cash)

            rec_type = row[0].value
            rec_type_list.append(rec_type)

        rec_dict = {'Amount': rec_list}
        rec_dataframe = pd.DataFrame(rec_dict, index=rec_type_list)

        sub_list = []
        sub_type_list = []


        for row in records.iter_rows(min_row=71, max_row=77):     # subscriptions expenses
            cash = row[ro].value        # the cash
            sub_list.append(cash)

            sub_type = row[0].value
            sub_type_list.append(sub_type)

        sub_dict = {'Amount': sub_list}
        sub_dataframe = pd.DataFrame(sub_dict, index=sub_type_list)

        pers_list = []
        pers_type_list = []

        for row in records.iter_rows(min_row=81, max_row=85):     # personal expenses
            cash = row[ro].value        # the cash
            pers_list.append(cash)

            pers_type = row[0].value
            pers_type_list.append(pers_type)

        pers_dict = {'Amount': pers_list}
        pers_dataframe = pd.DataFrame(pers_dict, index=pers_type_list)

        ob_list = []
        ob_type_list = []

        
        for row in records.iter_rows(min_row=89, max_row=93):     # obligations expenses
            cash = row[ro].value        # the cash
            ob_list.append(cash)

            ob_type = row[0].value
            ob_type_list.append(ob_type)

        ob_dict = {'Amount': ob_list}
        ob_dataframe = pd.DataFrame(ob_dict, index=ob_type_list)

        misc_list = []
        misc_type_list = []

        for row in records.iter_rows(min_row=97, max_row=101):     # misc expenses
            cash = row[ro].value        # the cash
            misc_list.append(cash)

            misc_type = row[0].value
            misc_type_list.append(misc_type)

        misc_dict = {'Amount': misc_list}
        misc_dataframe = pd.DataFrame(misc_dict, index=misc_type_list)

        total_list = []
        total_type_list = []

        for row in records.iter_rows(min_row=105, max_row=106):     # total expenses
            cash = row[ro].value        # the cash
            total_list.append(cash)

            total_type = row[0].value
            total_type_list.append(total_type)

        choice = int(input("""
                            Choose a category

                                [1] - Income
                                [2] - Home
                                [3] - Daily
                                [4] - Transport
                                [5] - Entertainment
                                [6] - Vacation
                                [7] - Reacreation
                                [8] - Subscription
                                [9] - Personal
                                [10] - Obligations
                                [11] - Misc
                    """))

        if choice == 1:      # Income
            print(income_dataframe)
        elif choice == 2:    # Home
            print(home_dataframe)
        elif choice == 3:    # Daily
            print(daily_dataframe)
        elif choice == 4:    # Trans
            print(trans_dataframe)
        elif choice == 5:    # Entertain
            print(entertain_dataframe)
        elif choice == 6:    # Vaca
            print(vaca_dataframe)
        elif choice == 7:    # Rec
            print(rec_dataframe)
        elif choice == 8:    # Sub
            print(sub_dataframe)
        elif choice == 9:    # Pers
            print(pers_dataframe)
        elif choice == 10:   # Ob
            print(ob_dataframe)
        elif choice == 11:   # Misc
            print(misc_dataframe.tr)


class year():           # whole year
    def total():
        print("not yet available")

# month.monthly_category()
# month.monthly_total()